#!/usr/bin/python

# Copyright 2017, Gurobi Optimization, Inc.

# Solve a multi-commodity flow problem.  Two products ('Pencils' and 'Pens')
# are produced in 2 cities ('Detroit' and 'Denver') and must be sent to
# warehouses in 3 cities ('Boston', 'New York', and 'Seattle') to
# satisfy demand ('inflow[h,i]').
#
# Flows on the transportation network must respect arc capacity constraints
# ('capacity[i,j]'). The objective is to minimize the sum of the arc
# transportation costs ('cost[i,j]').

from gurobipy import *
import random
import os
from openpyxl import Workbook

# 创建一个新的工作簿
wb = Workbook()

# 激活默认的工作表
ws = wb.active

ws['A1'] = 'Gen'
ws['B1'] = 'Ori'
outFile = open('record-real-min-cost.txt', 'w')
def f(file_path_gen, file_path_ori):
    # Model data
    print(file_path_ori)
    outFile.write(file_path_ori)
    outFile.write("\n")
    print(file_path_gen)
    outFile.write(file_path_gen)
    outFile.write("\n")
    commodities = ['Pencils', 'Pens']
    nodes = set()
    arcs = {}
    cost = {}
    inflow = {}
    with open(file_path_gen, 'r') as file:
        # 读取第一行，获取矩阵大小n
        content = file.read()
        lines = content.splitlines()
        for line in lines:
            if line == lines[0]:
                n = int(line)
            else:
                edge = line.split()
                nodes.add(edge[0])
                nodes.add(edge[1])
                arcs[edge[0], edge[1]] = random.randint(2000000,3000000)
                for commodity in commodities:
                    cost[commodity, edge[0], edge[1]] = random.randint(10, 30)
    with open(file_path_ori, 'r') as file:
        # 读取第一行，获取矩阵大小n
        content = file.read()
        lines = content.splitlines()
        for line in lines:
            if line != lines[0]:
                edge = line.split()
                for commodity in commodities:
                    x = random.randint(0, 5)
                    if (commodity,edge[0]) in inflow:
                        inflow[commodity, edge[0]] += x
                    else:
                        inflow[commodity, edge[0]] = x
                    if (commodity,edge[1]) in inflow:
                        inflow[commodity, edge[1]] -= x
                    else:
                        inflow[commodity, edge[1]] = -x
    arcs, capacity = multidict(arcs)
    #Create optimization model
    m = Model('min_cost')
    # Create variables
    # min-cost
    flow = m.addVars(commodities, arcs, obj = cost, vtype = GRB.INTEGER, name = "flow")
    # Arc capacity constraints
    m.addConstrs(
        (flow.sum('*', i, j) <= capacity[i, j] for i, j in arcs), "cap")
    # Flow conservation constraints
    m.addConstrs(
        (
            flow.sum(h, '*', j) + (inflow[h, j] if (h, j) in inflow else 0) == flow.sum(h, j, '*')
            for h in commodities
            for j in nodes
        ),
        "node")
    m.optimize()

    print(m.IsMIP)
    outFile.write("run time: ")
    outFile.write(str(m.Runtime))
    outFile.write("\n")
    '''
    # Print solution
    if m.status == GRB.Status.OPTIMAL:
        solution = m.getAttr('x', flow)
        for h in commodities:
            print('\nOptimal flows for %s:' % h)
            for i, j in arcs:
                if solution[h, i, j] > 0:
                    print('%s -> %s: %g' % (i, j, solution[h, i, j]))
    '''
    return m.Runtime


path = "/Users/songtianqi/vscode/data_real/graph3/"
files = os.listdir(path)
if ".DS_Store" in files:
    files.remove(".DS_Store")
files = sorted(files)
print(files)

m = len(files)
for i in range(0, 36):
    print(files[3 * i], files[3 * i + 1], files[3 * i +2])
    x1 = f(path + files[3 * i + 1], path + files[3 * i + 2])
    x2 = f(path + files[3 * i], path + files[3 * i + 2])
    x3 = f(path + files[3 * i + 2], path + files[3 * i + 2])
    ws.append([x1, x2, x3])
wb.save("result-real-min-cost.xlsx")