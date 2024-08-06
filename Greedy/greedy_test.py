# %%
import copy
import numpy as np
import os
from openpyxl import Workbook

def write_to_file(filename, data, mode='w', encoding='utf-8'):
    """
    将数据写入到文本文件。

    参数:
    filename (str): 文件名
    data (str): 要写入的数据
    mode (str): 文件打开模式，默认 'w'（写入模式），可以改为 'a'（追加模式）
    encoding (str): 文件编码，默认 'utf-8'
    """
    try:
        with open(filename, mode, encoding=encoding) as file:
            file.write(data)
    except Exception as e:
        print(f"写入文件时发生错误: {e}")
def create_seq(u, r_rev):
    seq = []
    # early_stop = 20
    while len(seq) < n + l:
        removed_a = set()
        # if exist a reachable to all rest B
        # pick all such a
        for item in u.items():
            if not item[1]:
                a = item[0]
                seq.append(a)
                removed_a.add(a)
                # print(a)
            else:
                break
        for a in removed_a:
            u.pop(a)
        # print(u)

        # update reachablity
        if removed_a:
            for value in r_rev.values():
                value -= set(removed_a)
            # print(r_rev)
            r_rev = dict(sorted(r_rev.items(), key = lambda item: len(item[1])))

        # if does not exist a reachable to all rest B
        # pick b from B that has the least reachable from rest A
        for item in r_rev.items():
            b = item[0]
            seq.append(b)
            r_rev.pop(b)
            # print(b)
            # update reachablity
            for value in u.values():
                value -= {b}
            # print(u)
            u = dict(sorted(u.items(), key = lambda item: len(item[1])))
            break
    return seq


def check_valid(seq, unreachablity):
    # check if there is seq[i] is unreachable to seq[j]
    for i in range(len(seq)):
        for j in range(i+1, len(seq)):
            if seq[i] in unreachablity and seq[j] in unreachablity[seq[i]]:
                return False
    return True


def check_lost_reachablity(seq, reverse_reachablity, outFile):
    lost_reachablity = []
    for i in range(len(seq)):
        for j in range(i):
            # check if there is seq[i] is reachable to seq[j]
            if seq[j] in reverse_reachablity and seq[i] in reverse_reachablity[seq[j]]:
                lost_reachablity.append((seq[i], seq[j]))
                write_to_file(outFile, str("s" + str(seq[i]) + " " + "t" + str(seq[j] - n)) + "\n", mode='a')
    return lost_reachablity

'''
def evaluate_seq(seq, reverse_reachablity):
    num_arc = 0
    prev_aux_status = None
    for i in range(len(seq)):
        # if seq[i] belongs to A
        if seq[i] < n:
            for j in range(i+1, len(seq)):
                # check whether there exist reachable seq[j] in the rest sequence
                if seq[j] in reverse_reachablity and seq[i] in reverse_reachablity[seq[j]]:
                    num_arc += 1
                if prev_aux_status == "out":
                    num_arc += 1
                prev_aux_status = "in"
                break
        # if seq[i] belongs to B
        else:
            for j in range(i):
                # check whether there exist seq[j] reachable to seq[i] in the past sequence
                if seq[i] in reverse_reachablity and seq[j] in reverse_reachablity[seq[i]]:
                    num_arc += 1
                prev_aux_status = "out"
                break
    # add arcs to fix lost reachability
    num_arc += len(check_lost_reachablity(seq, reverse_reachablity))
    return num_arc
'''
def evaluate_seq(seq, reverse_reachablity, outFile):
    m = l
    num_arc = 0
    write_to_file(outFile, str(n) + "\n")
    prev_aux_status = None
    for i in range(len(seq)):
        if seq[i] < n:
            for j in range(i+1, len(seq)):
                if seq[j] in reverse_reachablity and seq[i] in reverse_reachablity[seq[j]]:
                    num_arc += 1
                    if prev_aux_status == "out":
                        num_arc += 1
                        m = m + 1
                        if m > n:
                            write_to_file(outFile, str("t" + str(m - 1) + " " + "t" + str(m)) + "\n",mode= 'a')
                    write_to_file(outFile, str("s" + str(seq[i]) + " " + "t" + str(m)) + "\n",mode= 'a')
                    prev_aux_status = "in"
                    break
        else:
            for j in range(i):
                if seq[i] in reverse_reachablity and seq[j] in reverse_reachablity[seq[i]]:
                    num_arc += 1
                    write_to_file(outFile, str("t" + str(m) + " " + "t" + str(seq[i] - n)) + "\n", mode = 'a')
                    prev_aux_status = "out"
                    break
    num_arc += len(check_lost_reachablity(seq, reverse_reachablity, outFile))
    return num_arc, m - l

def read_matrix_from_file(file_path):
    with open(file_path, 'r') as file:
        # 读取第一行，获取矩阵大小n
        n = int(file.readline().strip())
        # 读取后续n行，构建矩阵
        matrix = []
        for _ in range(n):
            row = list(map(int, file.readline().strip().split()))
            matrix.append(row)
    return np.array(matrix)  # 将列表转换为numpy数组并返回
# %%
path = "/Users/songtianqi/vscode/data_real/data_repaired"
path_out = "/Users/songtianqi/vscode/data_real/data_repaired/graph_greedy"
files = os.listdir(path)
files = sorted(files)

# 创建一个新的工作簿
wb = Workbook()

# 激活默认的工作表
ws = wb.active

ws['A1'] = 'Source'
ws['B1'] = 'Size'
ws['C1'] = '原图边数'
ws['D1'] = '生成图边数'
ws['E1'] = '生成的结点数'
for file in files:
    if file[0] != "d":
        continue
    print(file)
    outFile = path_out + "/" + file[:-4] + "_gen_greedy.txt"
    p_reachable = read_matrix_from_file(path+ "/" +file)
    n = len(p_reachable)
    l = len(p_reachable[0])
    unreachablity = {
        i: {j + n for j in range(l) if p_reachable[i, j] == 0} for i in range(n)}
    reverse_reachablity = {
        j + n: {i for i in range(n) if p_reachable[i, j] == 1} for j in range(l)}
    u = copy.deepcopy(unreachablity)
    u = dict(sorted(u.items(), key=lambda item: len(item[1])))
    r_rev = copy.deepcopy(reverse_reachablity)
    r_rev = dict(sorted(r_rev.items(), key=lambda item: len(item[1])))

    seq = create_seq(u, r_rev)
    for i in range(len(seq)):
        if(seq[i] < n):
            print('a', end="")
        else:
            print('b', end="")
    print(" ")
    # print(f"Legality: {check_valid(seq, unreachablity)}")
    cnt_arc, cnt_dot = evaluate_seq(seq, reverse_reachablity, outFile)
    print(f"num arcs created by heuristic:" + str(cnt_arc))
    print(f"reachability size: {(p_reachable == 1).sum()}")
    ws.append([file, n,(p_reachable == 1).sum(), cnt_arc, cnt_dot])

wb.save('greedy_result_real_data_repair.xlsx')
# %%
